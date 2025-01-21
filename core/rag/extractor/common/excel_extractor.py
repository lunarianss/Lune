import logging
import pandas as pd
import os
from domain.rag.entity import document
from typing import Optional
from core.rag.extractor.interface import extractor_base
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelExtractor(extractor_base.BaseExtractor):

    def __init__(self, file_path: str, encoding: Optional[str] = None, autodetect_encoding: bool = False):
        self._file_path = file_path
        self._encoding = encoding
        self._autodetect_encoding = autodetect_encoding

    def extract(self):
        documents = []
        file_extension = os.path.splitext(self._file_path)[-1].lower()

        if file_extension == ".xlsx":
            wb = load_workbook(self._file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = sheet.values

                print(data, "data values")

                try:
                    cols = next(data)
                except StopIteration:
                    continue

                df = pd.DataFrame(data, columns=cols)
                df.dropna(how="all", inplace=True)

                for index, row in df.iterrows():
                    page_content = []
                    for col_index, (k, v) in enumerate(row.items()):
                        if not pd.notna(v):
                            continue
                        cell = sheet.cell(
                            row=index + 2, column=col_index + 1
                        )

                        if cell.hyperlink:
                            value = f"[{v}]({cell.hyperlink.target})"
                            page_content.append(f'"{k}":"{value}"')
                        else:
                            page_content.append(f'"{k}":"{v}"')

                    documents.append(
                        document.Document(page_content=";".join(page_content),
                                          metadata={"source": self._file_path})
                    )
        elif file_extension == ".xls":
            excel_file = pd.ExcelFile(self._file_path, engine="xlrd")

            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=sheet_name)
                df.dropna(how="all", inplace=True)

                for _, row in df.iterrows():
                    page_content = []
                    for k, v in row.items():
                        if pd.notna(v):
                            page_content.append(f'"{k}":"{v}"')
                    documents.append(
                        document.Document(page_content=";".join(page_content),
                                          metadata={"source": self._file_path})
                    )
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        return documents
